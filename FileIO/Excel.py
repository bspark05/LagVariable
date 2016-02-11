#-*- coding: utf-8 -*-

import xlrd
import openpyxl
import xlwt
from openpyxl import Workbook


def excelRead(filepath, sheetname):
    workbook = xlrd.open_workbook(filepath)
    worksheet = workbook.sheet_by_name(sheetname)
    
    num_rows = worksheet.nrows -1
    curr_row = -1
    result = []
    
    while curr_row < num_rows:
        curr_row += 1
        row = worksheet.row(curr_row)
        result.append(row)
    
    return result


# first row - data
def excelWriteOnExistingFile(filepath, sheetname, columnNum, insert): 
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active
    
    num_rows = ws.nrows -1
    curr_row = -1
    
    while curr_row < num_rows:
        curr_row += 1
        
        worksheet[columnNum+str(curr_row+1)] = insert[curr_row][0]
        asciiNum = ord(columnNum)+1
        columnNumPlus = chr(asciiNum)
        worksheet[columnNumPlus+str(curr_row+1)] = insert[curr_row][1]
        
    workbook.save(filepath)
    print('saved successfully in existing file!')
    
    
    # .xlsx file available
def excelWriteOnExistingFile2(filepath, sheetname, insert):
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active
    
    i1 = 0 
    i2 = ord('A') 
    num_rows = ws.nrows
    
    
    while i1<len(insert[0]):
        j=0
        k=j+1
        while j<len(insert): #2666
            #print(type(insertList[j][i1]))
            if str(type(insert[j][i1])) == "<class 'xlrd.sheet.Cell'>":
                worksheet[chr(i2)+str(k+num_rows)] = insert[j][i1].value
            else:
                worksheet[chr(i2)+str(k+num_rows)] = insert[j][i1].encode('utf-8')
            j+=1
            k+=1
        i1+=1
        i2+=1
        
    workbook.save(filepath)
    print('saved successfully!')

#first row - field name    
def excelWriteOnExistingFile3(filepath, sheetname, columnNum, insert): 
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active
    
    num_rows = ws.nrows -1
    curr_row = -1
    
    while curr_row < num_rows -1:
        curr_row += 1
        
        worksheet[columnNum+str(curr_row+2)] = insert[curr_row]
                
    workbook.save(filepath)
    print('saved successfully in existing file!')
    
def excelWriteNewFile(filepath, sheetname, insertList):
    '''
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheetname)
    
    i=0
    while i<len(insert[0]):
        j=0
        while j<len(insert):
            print(insert[j][i])
            #print(type(insert[j][i]))
            ws.write(j, i, unicode(insert[j][i]))
            j+=1
        i+=1
        
    wb.save(filepath)
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    
    i1 = 0
    i2 = ord('A')
    while i1<len(insertList[0]):
        j=0
        k=j+1
        while j<len(insertList):
            #print(type(insertList[j][i1]))
            if str(type(insertList[j][i1])) == "<class 'xlrd.sheet.Cell'>":
                ws[chr(i2)+str(k)] = insertList[j][i1].value
            else:
                ws[chr(i2)+str(k)] = insertList[j][i1].encode('utf-8')
            j+=1
            k+=1
        i1+=1
        i2+=1
        
    wb.save(filepath)
    print('saved successfully in a new file!')
    
def xlsToXlsx(filepath, sheetname):
    xlsList = excelRead(filepath, sheetname)
    
    resultList = engToKor(filepath, sheetname)
    engFilepath = resultList[0]+'.xlsx'
    engSheetname = resultList[1]
    
    excelWriteNewFile(engFilepath, engSheetname, xlsList)
    
    fileInfoList = [str(engFilepath), engSheetname]
    return fileInfoList 
    
def engToKor(filepath, sheetname):
    sale = '매매'
    rent = '전월'
    
    # remove last 4 chars (.xls)
    filepath = filepath[:-4]
    
    firstPart = filepath[0:6]
    secondPart = filepath[6:8]
    thirdPart1 = filepath[8:]
    thirdPart2 = filepath[9:]
    
    
    def ifType(thirdPart):
        apartment = '아파트'
        detached = '단독_다가구'
        tenement = '연립_다세대'
        
        if thirdPart == apartment.decode('utf-8'):
            thirdPart = 'Apartment'
        elif thirdPart == detached.decode('utf-8'):
            thirdPart = 'Detached'
        elif thirdPart == tenement.decode('utf-8'):
            thirdPart = 'Tenement'
        return thirdPart
    
    if secondPart == sale.decode('utf-8'):
        secondPart = 'Sale'
        thirdPart = ifType(thirdPart1)
    elif secondPart == rent.decode('utf-8'):
        secondPart = 'Rent'
        thirdPart = ifType(thirdPart2)
    
    resultFilename = firstPart+secondPart+thirdPart
    
    seoul = '서울'
    busan = '부산'
    if sheetname == seoul.decode('utf-8'):
        resultSheetname = 'Seoul'
    elif sheetname == busan.decode('utf-8'):
        resultSheetname = 'Busan'
            
    return [resultFilename, resultSheetname] 

    
    
    